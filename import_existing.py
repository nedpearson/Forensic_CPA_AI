"""
Import existing analysis data from the Excel files into the forensic auditor database.
Imports:
  1. Bank_Analysis.xlsx - All 319 bank transactions from Account #130109
  2. FINAL_ANALYSIS_NO_ELI.xlsx - Venmo details, Green Card, Silver Card data
"""
import os
import sys
import openpyxl
from database import (
    init_db, get_db, get_or_create_account, add_document, add_transaction
)
from categorizer import categorize_transaction

BASE_DIR = os.path.dirname(os.path.dirname(__file__))


def import_bank_transactions():
    """Import from Bank_Analysis.xlsx -> 'All Transactions' sheet."""
    filepath = os.path.join(BASE_DIR, 'Bank_Analysis.xlsx')
    if not os.path.exists(filepath):
        print(f"  SKIP: {filepath} not found")
        return 0

    wb = openpyxl.load_workbook(filepath)
    ws = wb['All Transactions']

    # Create account
    account_id = get_or_create_account(
        account_name='Gulf Coast Recovery - Checking',
        account_number='130109',
        account_type='bank',
        institution='Bank of St. Francisville',
        cardholder_name='Gulf Coast Recovery of Baton Rouge, LLC'
    )

    # Create document record
    doc_id = add_document(
        filename='Bank_Analysis.xlsx',
        original_path=filepath,
        file_type='xlsx',
        doc_category='bank_statement',
        account_id=account_id,
        statement_start='2024-08-01',
        statement_end='2026-01-30',
        notes='Imported from existing analysis spreadsheet'
    )

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_str, description, withdrawal, deposit, category, payment_method = row[:6]
        if not description:
            continue

        # Determine amount (negative for withdrawals, positive for deposits)
        amount = 0
        trans_type = 'debit'
        if deposit and float(deposit) > 0:
            amount = float(deposit)
            trans_type = 'deposit'
        elif withdrawal and float(withdrawal) > 0:
            amount = -float(withdrawal)
            trans_type = 'debit'

        if amount == 0 and not deposit and not withdrawal:
            continue

        # Normalize date to YYYY-MM-DD ish format
        date_normalized = normalize_date(str(date_str) if date_str else '')

        # Map old categories to new ones
        cat_result = categorize_transaction(description, amount, trans_type, payment_method or '')

        # Override with existing category mapping
        old_cat = str(category or '').upper()
        if old_cat == 'DEPOSIT':
            cat_result['category'] = 'Deposits'
        elif old_cat == 'TRANSFER':
            cat_result['is_transfer'] = 1
            if 'Transfers' not in cat_result['category'] and cat_result['category'] == 'Uncategorized':
                cat_result['category'] = 'Transfers Out' if amount < 0 else 'Transfers In'
        elif old_cat == 'PERSONAL':
            cat_result['is_personal'] = 1
            if 'Personal' not in cat_result['category']:
                cat_result['category'] = 'Personal - Other'
        elif old_cat in ('BUSINESS_LARGE', 'BUSINESS_MEDIUM', 'BUSINESS_SMALL'):
            cat_result['is_business'] = 1
            if 'Business' not in cat_result['category']:
                cat_result['category'] = 'Business - Other'
        elif old_cat == 'CHECK':
            cat_result['category'] = 'Check Payment'
        elif old_cat == 'FEE':
            cat_result['category'] = 'Fees - NSF/Overdraft'

        add_transaction(
            doc_id=doc_id,
            account_id=account_id,
            trans_date=date_normalized,
            post_date=date_normalized,
            description=str(description),
            amount=amount,
            trans_type=trans_type,
            category=cat_result['category'],
            subcategory=cat_result.get('subcategory'),
            cardholder_name='',
            card_last_four='0109',
            payment_method=str(payment_method or '').lower(),
            is_transfer=cat_result.get('is_transfer', 0),
            is_personal=cat_result.get('is_personal', 0),
            is_business=cat_result.get('is_business', 0),
            is_flagged=cat_result.get('is_flagged', 0),
            flag_reason=cat_result.get('flag_reason'),
        )
        count += 1

    print(f"  Imported {count} bank transactions from Bank_Analysis.xlsx")
    return count


def import_venmo_transactions():
    """Import from FINAL_ANALYSIS_NO_ELI.xlsx -> 'All Venmo Detail' sheet."""
    filepath = os.path.join(BASE_DIR, '..', 'FINAL_ANALYSIS_NO_ELI.xlsx')
    if not os.path.exists(filepath):
        print(f"  SKIP: {filepath} not found")
        return 0

    wb = openpyxl.load_workbook(filepath)
    ws = wb['All Venmo Detail']

    account_id = get_or_create_account(
        account_name='Venmo',
        account_number='venmo',
        account_type='venmo',
        institution='Venmo',
    )

    doc_id = add_document(
        filename='FINAL_ANALYSIS_NO_ELI.xlsx (Venmo)',
        original_path=filepath,
        file_type='xlsx',
        doc_category='venmo',
        account_id=account_id,
        statement_start='2024-02-01',
        statement_end='2026-02-28',
        notes='Venmo transactions imported from final analysis (Eli excluded)'
    )

    count = 0
    # Find header row (row 4 based on our inspection)
    for row in ws.iter_rows(min_row=5, values_only=True):
        batch, month, date, recipient, amount, funding_source, note, category, vid, status = row[:10]
        if not recipient or not amount:
            continue

        try:
            amt = float(amount)
        except (ValueError, TypeError):
            continue

        date_str = str(date) if date else ''
        # Normalize date
        date_normalized = normalize_date(date_str)

        desc = f"VENMO: {recipient}"
        if note:
            desc += f" - {note}"
        if funding_source:
            desc += f" (via {str(funding_source)[:30]})"

        cat_result = categorize_transaction(desc, -amt, 'debit', 'venmo')
        cat_result['category'] = f"Venmo - Payment"

        # Check if it's a James payment
        is_flagged = 0
        flag_reason = None
        recipient_upper = str(recipient).upper()
        if 'JAMES' in recipient_upper:
            is_flagged = 1
            flag_reason = f"Venmo payment to JAMES: ${amt:.2f}"

        add_transaction(
            doc_id=doc_id,
            account_id=account_id,
            trans_date=date_normalized,
            post_date=date_normalized,
            description=desc,
            amount=-abs(amt),
            trans_type='debit',
            category=cat_result['category'],
            subcategory=str(category or ''),
            cardholder_name=str(recipient or ''),
            card_last_four='',
            payment_method='venmo',
            is_transfer=1,
            is_personal=0,
            is_business=0,
            is_flagged=is_flagged,
            flag_reason=flag_reason,
            transfer_to_account=str(recipient or ''),
        )
        count += 1

    print(f"  Imported {count} Venmo transactions")
    return count


def import_green_card():
    """Import from FINAL_ANALYSIS_NO_ELI.xlsx -> 'Green Card 5762 Detail'."""
    filepath = os.path.join(BASE_DIR, '..', 'FINAL_ANALYSIS_NO_ELI.xlsx')
    if not os.path.exists(filepath):
        return 0

    wb = openpyxl.load_workbook(filepath)
    ws = wb['Green Card 5762 Detail']

    account_id = get_or_create_account(
        account_name='Capital One Spark Cash Plus (Green)',
        account_number='5762',
        account_type='credit_card',
        institution='Capital One',
        cardholder_name='Gerald T Pearson',
        card_last_four='5762'
    )

    # Also add James's sub-card
    james_account_id = get_or_create_account(
        account_name='Capital One Spark Cash Plus (Green) - James',
        account_number='9719',
        account_type='credit_card',
        institution='Capital One',
        cardholder_name='James Hendrick',
        card_last_four='9719'
    )

    doc_id = add_document(
        filename='FINAL_ANALYSIS_NO_ELI.xlsx (Green Card)',
        original_path=filepath,
        file_type='xlsx',
        doc_category='credit_card',
        account_id=account_id,
        statement_start='2024-07-01',
        statement_end='2025-12-31',
        notes='Green Card (5762) transactions - Gerald Pearson & James Hendrick'
    )

    count = 0
    for row in ws.iter_rows(min_row=6, values_only=True):
        quarter, month, date, card, merchant, category, amount, notes = (list(row) + [None]*8)[:8]
        if not merchant or not amount:
            continue

        try:
            amt = float(amount)
        except (ValueError, TypeError):
            continue

        date_str = normalize_date(str(date) if date else str(month or ''))
        card_str = str(card or '')

        # Determine cardholder
        cardholder = 'Gerald T Pearson'
        acct_id = account_id
        card_four = '5762'
        if '9719' in card_str or 'JAMES' in card_str.upper() or 'HENDRICK' in card_str.upper():
            cardholder = 'James Hendrick'
            acct_id = james_account_id
            card_four = '9719'

        cat_result = categorize_transaction(str(merchant), -amt, 'debit', 'credit')
        category_str = str(category or '')

        # Map categories
        if 'FEE' in category_str.upper():
            cat_result['category'] = 'Fees - Late Payment'
        elif 'DINING' in category_str.upper() or 'FOOD' in category_str.upper() or 'RESTAURANT' in category_str.upper():
            cat_result['category'] = 'Personal - Dining'
            cat_result['is_personal'] = 1
        elif 'GAS' in category_str.upper() or 'FUEL' in category_str.upper():
            cat_result['category'] = 'Business - Supplies'
            cat_result['is_business'] = 1
        elif cat_result['category'] == 'Uncategorized':
            cat_result['category'] = 'Business - Other'
            cat_result['is_business'] = 1

        add_transaction(
            doc_id=doc_id,
            account_id=acct_id,
            trans_date=date_str,
            post_date=date_str,
            description=str(merchant),
            amount=-abs(amt),
            trans_type='debit',
            category=cat_result['category'],
            subcategory=category_str,
            cardholder_name=cardholder,
            card_last_four=card_four,
            payment_method='credit',
            is_transfer=cat_result.get('is_transfer', 0),
            is_personal=cat_result.get('is_personal', 0),
            is_business=cat_result.get('is_business', 0),
            is_flagged=cat_result.get('is_flagged', 0),
            flag_reason=cat_result.get('flag_reason'),
        )
        count += 1

    print(f"  Imported {count} Green Card transactions")
    return count


def import_silver_card():
    """Import from FINAL_ANALYSIS_NO_ELI.xlsx -> 'Silver Card 6248 Detail'."""
    filepath = os.path.join(BASE_DIR, '..', 'FINAL_ANALYSIS_NO_ELI.xlsx')
    if not os.path.exists(filepath):
        return 0

    wb = openpyxl.load_workbook(filepath)
    ws = wb['Silver Card 6248 Detail']

    account_id = get_or_create_account(
        account_name='Capital One Silver (6248)',
        account_number='6248',
        account_type='credit_card',
        institution='Capital One',
        cardholder_name='Gerald T Pearson',
        card_last_four='6248'
    )

    doc_id = add_document(
        filename='FINAL_ANALYSIS_NO_ELI.xlsx (Silver Card)',
        original_path=filepath,
        file_type='xlsx',
        doc_category='credit_card',
        account_id=account_id,
        statement_start='2024-01-01',
        statement_end='2026-02-28',
        notes='Silver Card (6248) monthly summaries'
    )

    count = 0
    for row in ws.iter_rows(min_row=6, values_only=True):
        vals = list(row) + [None]*7
        month, spending, fees, interest, status, top_categories, notes = vals[:7]
        if not month or not spending:
            continue

        try:
            spent = float(spending)
        except (ValueError, TypeError):
            continue

        date_str = normalize_date(str(month))

        # Add the spending as a summary transaction
        if spent > 0:
            add_transaction(
                doc_id=doc_id, account_id=account_id,
                trans_date=date_str, post_date=date_str,
                description=f"Silver Card Monthly Spending ({month}) - {top_categories or 'Various'}",
                amount=-abs(spent), trans_type='debit',
                category='Business - Other',
                cardholder_name='Gerald T Pearson', card_last_four='6248',
                payment_method='credit', is_business=1,
                user_notes=str(notes or ''),
            )
            count += 1

        # Add fees if present
        try:
            fee_amt = float(fees) if fees else 0
        except (ValueError, TypeError):
            fee_amt = 0
        if fee_amt > 0:
            add_transaction(
                doc_id=doc_id, account_id=account_id,
                trans_date=date_str, post_date=date_str,
                description=f"Silver Card Fees ({month})",
                amount=-abs(fee_amt), trans_type='fee',
                category='Fees - Late Payment',
                cardholder_name='Gerald T Pearson', card_last_four='6248',
                payment_method='credit',
            )
            count += 1

        # Add interest if present
        try:
            int_amt = float(interest) if interest else 0
        except (ValueError, TypeError):
            int_amt = 0
        if int_amt > 0:
            add_transaction(
                doc_id=doc_id, account_id=account_id,
                trans_date=date_str, post_date=date_str,
                description=f"Silver Card Interest ({month})",
                amount=-abs(int_amt), trans_type='fee',
                category='Fees - Service Charge',
                cardholder_name='Gerald T Pearson', card_last_four='6248',
                payment_method='credit',
            )
            count += 1

    print(f"  Imported {count} Silver Card entries")
    return count


def normalize_date(date_str):
    """Best-effort date normalization to YYYY-MM-DD."""
    if not date_str or date_str == 'None':
        return ''

    date_str = date_str.strip().rstrip('.')

    # Already in YYYY-MM-DD format
    if len(date_str) >= 10 and date_str[4] == '-':
        return date_str[:10]

    # Handle "2024-02-01 00:00:00" format
    if ' ' in date_str and '-' in date_str:
        return date_str.split(' ')[0]

    # Month abbreviation + day: "Aug 01", "Sep 03."
    import re
    month_map = {
        'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04',
        'may': '05', 'jun': '06', 'jul': '07', 'aug': '08',
        'sep': '09', 'oct': '10', 'nov': '11', 'dec': '12'
    }

    match = re.match(r'(\w{3})\s*(\d{1,2})', date_str)
    if match:
        mon = match.group(1).lower()
        day = match.group(2).zfill(2)
        if mon in month_map:
            # Guess year based on context (most data is 2024-2026)
            return f"2025-{month_map[mon]}-{day}"

    # "Month Year" like "January 2025"
    full_months = {
        'january': '01', 'february': '02', 'march': '03', 'april': '04',
        'may': '05', 'june': '06', 'july': '07', 'august': '08',
        'september': '09', 'october': '10', 'november': '11', 'december': '12'
    }
    for m, num in full_months.items():
        if m in date_str.lower():
            year_match = re.search(r'(\d{4})', date_str)
            yr = year_match.group(1) if year_match else '2025'
            return f"{yr}-{num}-01"

    # "YYYY_MM" or similar
    match = re.match(r'(\d{4})[_/-](\d{1,2})', date_str)
    if match:
        return f"{match.group(1)}-{match.group(2).zfill(2)}-01"

    return date_str


def main():
    print("=" * 60)
    print("IMPORTING EXISTING DATA INTO FORENSIC AUDITOR")
    print("=" * 60)

    # Clear existing data first
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) as cnt FROM transactions")
    existing = cursor.fetchone()['cnt']
    if existing > 0:
        print(f"\n  Database already has {existing} transactions.")
        response = input("  Clear and re-import? (yes/no): ").strip().lower()
        if response != 'yes':
            print("  Skipping import.")
            conn.close()
            return
        cursor.execute("DELETE FROM transactions")
        cursor.execute("DELETE FROM documents")
        cursor.execute("DELETE FROM accounts")
        cursor.execute("DELETE FROM audit_log")
        conn.commit()
        print("  Cleared existing data.")
    conn.close()

    total = 0

    print("\n1. Importing Bank of St. Francisville transactions...")
    total += import_bank_transactions()

    print("\n2. Importing Venmo transactions...")
    total += import_venmo_transactions()

    print("\n3. Importing Green Card (5762) transactions...")
    total += import_green_card()

    print("\n4. Importing Silver Card (6248) transactions...")
    total += import_silver_card()

    print(f"\n{'=' * 60}")
    print(f"IMPORT COMPLETE: {total} total transactions loaded")
    print(f"{'=' * 60}")
    print(f"\nRun the app: python app.py")
    print(f"Then open: http://localhost:5000")


if __name__ == '__main__':
    init_db()
    main()
